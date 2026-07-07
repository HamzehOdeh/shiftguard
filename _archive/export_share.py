"""
export_share.py - Export, Print, and Sharing Module for Workforce Compliance AI

Handles all export (PDF, Excel, CSV, iCal), print formats (breakroom poster,
individual cards), sharing (links, email, SMS, Slack), import (CSV, Excel,
Kronos, ADP), and report generation (weekly summary, monthly executive).
"""

import csv
import hashlib
import json
import os
import secrets
import smtplib
import uuid
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import StringIO
from typing import Any, Dict, List, Optional

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import fpdf2 for PDF support
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False


# =============================================================================
# 1. SCHEDULE EXPORT
# =============================================================================

def export_to_pdf(schedule: Dict, filepath: str) -> str:
    """
    Export schedule to a printable PDF grid.

    Format: employees down the left, Mon-Sun across the top.
    Each cell shows shift time + role.
    Color coding: Tier 1 = blue, Tier 2 = green, Tier 3 = orange.
    Header: facility name, week dates, pool status summary.
    Footer: compliance status, total headcount per shift.

    Falls back to text-based output if fpdf2 is not installed.
    """
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    week_end = schedule.get("week_end", "2026-01-11")
    pool_status = schedule.get("pool_status", "Active - 85% fill rate")
    compliance_status = schedule.get("compliance_status", "Compliant - No violations")
    employees = schedule.get("employees", [])
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    if FPDF_AVAILABLE:
        pdf = FPDF(orientation="L", format="letter")
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Header
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, f"{facility} - Weekly Schedule", ln=True, align="C")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, f"Week: {week_start} to {week_end}", ln=True, align="C")
        pdf.cell(0, 6, f"Pool Status: {pool_status}", ln=True, align="C")
        pdf.ln(5)

        # Table header
        col_width_name = 40
        col_width_day = 32
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_width_name, 8, "Employee", border=1)
        for day in days:
            pdf.cell(col_width_day, 8, day, border=1, align="C")
        pdf.ln()

        # Color map for tiers
        tier_colors = {
            1: (173, 216, 230),   # Light blue
            2: (144, 238, 144),   # Light green
            3: (255, 200, 130),   # Light orange
        }

        # Employee rows
        pdf.set_font("Helvetica", "", 8)
        for emp in employees:
            pdf.cell(col_width_name, 8, emp.get("name", "Unknown"), border=1)
            shifts = emp.get("shifts", {})
            tier = emp.get("tier", 1)
            r, g, b = tier_colors.get(tier, (255, 255, 255))

            for day in days:
                shift = shifts.get(day, {})
                if shift:
                    text = f"{shift.get('start', '')}-{shift.get('end', '')} {shift.get('role', '')}"
                    pdf.set_fill_color(r, g, b)
                    pdf.cell(col_width_day, 8, text[:15], border=1, align="C", fill=True)
                else:
                    pdf.cell(col_width_day, 8, "OFF", border=1, align="C")
            pdf.ln()

        # Footer
        pdf.ln(5)
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, f"Compliance: {compliance_status}", ln=True)

        # Headcount per shift
        headcount = _calculate_headcount(schedule)
        headcount_str = " | ".join([f"{day}: {count}" for day, count in headcount.items()])
        pdf.cell(0, 6, f"Headcount: {headcount_str}", ln=True)

        pdf.output(filepath)
        return f"PDF exported to {filepath}"
    else:
        # Text-based fallback
        return _export_to_text_pdf(schedule, filepath)


def _export_to_text_pdf(schedule: Dict, filepath: str) -> str:
    """Text-based fallback when fpdf2 is not installed."""
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    week_end = schedule.get("week_end", "2026-01-11")
    pool_status = schedule.get("pool_status", "Active - 85% fill rate")
    compliance_status = schedule.get("compliance_status", "Compliant - No violations")
    employees = schedule.get("employees", [])
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    lines = []
    lines.append("=" * 100)
    lines.append(f"  {facility} - Weekly Schedule".center(100))
    lines.append(f"  Week: {week_start} to {week_end}".center(100))
    lines.append(f"  Pool Status: {pool_status}".center(100))
    lines.append("=" * 100)
    lines.append("")

    # Header row
    header = f"{'Employee':<20}" + "".join([f"{d:<12}" for d in days])
    lines.append(header)
    lines.append("-" * 100)

    # Employee rows
    for emp in employees:
        name = emp.get("name", "Unknown")[:18]
        shifts = emp.get("shifts", {})
        tier = emp.get("tier", 1)
        tier_marker = f"[T{tier}]"
        row = f"{name:<16}{tier_marker:<4}"
        for day in days:
            shift = shifts.get(day, {})
            if shift:
                cell = f"{shift.get('start', '')}-{shift.get('end', '')}".strip()
            else:
                cell = "OFF"
            row += f"{cell:<12}"
        lines.append(row)

    lines.append("-" * 100)
    lines.append(f"Compliance: {compliance_status}")
    headcount = _calculate_headcount(schedule)
    headcount_str = " | ".join([f"{day}: {count}" for day, count in headcount.items()])
    lines.append(f"Headcount: {headcount_str}")
    lines.append("=" * 100)

    # Write as .txt (text-based PDF fallback)
    txt_filepath = filepath.replace(".pdf", ".txt") if filepath.endswith(".pdf") else filepath + ".txt"
    with open(txt_filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return f"Text-based schedule exported to {txt_filepath} (fpdf2 not installed for PDF)"


def export_to_excel(schedule: Dict, filepath: str) -> str:
    """
    Export schedule to a full Excel workbook with 4 sheets:
      Sheet 1: Schedule grid (same layout as PDF)
      Sheet 2: Employee details (hours, tier, reliability score)
      Sheet 3: Cost summary (premiums paid, savings vs traditional)
      Sheet 4: Compliance report (violations, if any)
    """
    if not OPENPYXL_AVAILABLE:
        return "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = openpyxl.Workbook()
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    employees = schedule.get("employees", [])
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")

    # Tier fill colors
    tier_fills = {
        1: PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # Blue
        2: PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Green
        3: PatternFill(start_color="FFC882", end_color="FFC882", fill_type="solid"),  # Orange
    }
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Sheet 1: Schedule Grid ---
    ws1 = wb.active
    ws1.title = "Schedule"
    ws1["A1"] = f"{facility} - Week of {week_start}"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:H1")

    # Headers
    ws1["A3"] = "Employee"
    ws1["A3"].font = header_font
    for i, day in enumerate(days, start=2):
        cell = ws1.cell(row=3, column=i, value=day)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, emp in enumerate(employees, start=4):
        ws1.cell(row=row_idx, column=1, value=emp.get("name", "Unknown")).border = thin_border
        shifts = emp.get("shifts", {})
        tier = emp.get("tier", 1)
        for col_idx, day in enumerate(days, start=2):
            shift = shifts.get(day, {})
            if shift:
                text = f"{shift.get('start', '')}-{shift.get('end', '')} {shift.get('role', '')}"
            else:
                text = "OFF"
            cell = ws1.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if shift:
                cell.fill = tier_fills.get(tier, PatternFill())

    # Auto-width
    for col in range(1, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet 2: Employee Details ---
    ws2 = wb.create_sheet("Employee Details")
    detail_headers = ["Employee ID", "Name", "Tier", "Weekly Hours", "Reliability Score", "Role", "Rate ($/hr)"]
    for i, h in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.border = thin_border

    for row_idx, emp in enumerate(employees, start=2):
        ws2.cell(row=row_idx, column=1, value=emp.get("id", "")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=emp.get("name", "")).border = thin_border
        ws2.cell(row=row_idx, column=3, value=emp.get("tier", 1)).border = thin_border
        ws2.cell(row=row_idx, column=4, value=emp.get("weekly_hours", 40)).border = thin_border
        ws2.cell(row=row_idx, column=5, value=emp.get("reliability_score", 0.95)).border = thin_border
        ws2.cell(row=row_idx, column=6, value=emp.get("role", "")).border = thin_border
        ws2.cell(row=row_idx, column=7, value=emp.get("rate", 25.0)).border = thin_border

    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # --- Sheet 3: Cost Summary ---
    ws3 = wb.create_sheet("Cost Summary")
    cost_data = schedule.get("cost_summary", {})
    ws3["A1"] = "Cost Summary"
    ws3["A1"].font = title_font

    cost_rows = [
        ("Total Regular Pay", cost_data.get("regular_pay", 0)),
        ("Overtime Premiums", cost_data.get("overtime_premiums", 0)),
        ("Tier 2 Premiums (15%)", cost_data.get("tier2_premiums", 0)),
        ("Tier 3 Premiums (25%)", cost_data.get("tier3_premiums", 0)),
        ("Total Pool Cost", cost_data.get("total_pool_cost", 0)),
        ("Traditional Staffing Cost", cost_data.get("traditional_cost", 0)),
        ("Net Savings", cost_data.get("savings", 0)),
        ("Savings Percentage", f"{cost_data.get('savings_pct', 0):.1f}%"),
    ]
    for row_idx, (label, value) in enumerate(cost_rows, start=3):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=row_idx, column=2, value=value)

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 18

    # --- Sheet 4: Compliance Report ---
    ws4 = wb.create_sheet("Compliance")
    ws4["A1"] = "Compliance Report"
    ws4["A1"].font = title_font

    violations = schedule.get("violations", [])
    ws4.cell(row=3, column=1, value="Status").font = header_font
    ws4.cell(row=3, column=2, value="Compliant" if not violations else f"{len(violations)} Violation(s)")

    if violations:
        ws4.cell(row=5, column=1, value="Violation").font = header_font
        ws4.cell(row=5, column=2, value="Employee").font = header_font
        ws4.cell(row=5, column=3, value="Details").font = header_font
        for row_idx, v in enumerate(violations, start=6):
            ws4.cell(row=row_idx, column=1, value=v.get("type", ""))
            ws4.cell(row=row_idx, column=2, value=v.get("employee", ""))
            ws4.cell(row=row_idx, column=3, value=v.get("details", ""))

    for col in range(1, 4):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    wb.save(filepath)
    return f"Excel workbook exported to {filepath}"


def export_to_csv(schedule: Dict, filepath: str) -> str:
    """
    Export schedule to a flat CSV file.
    Columns: employee_id, name, date, start, end, role, tier, rate, shift_type
    """
    employees = schedule.get("employees", [])
    week_start_str = schedule.get("week_start", "2026-01-05")
    days_map = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6}

    try:
        week_start_date = datetime.strptime(week_start_str, "%Y-%m-%d")
    except ValueError:
        week_start_date = datetime(2026, 1, 5)

    rows = []
    for emp in employees:
        emp_id = emp.get("id", "")
        name = emp.get("name", "")
        tier = emp.get("tier", 1)
        rate = emp.get("rate", 25.0)
        shifts = emp.get("shifts", {})

        for day, offset in days_map.items():
            shift = shifts.get(day, {})
            if shift:
                date = (week_start_date + timedelta(days=offset)).strftime("%Y-%m-%d")
                rows.append({
                    "employee_id": emp_id,
                    "name": name,
                    "date": date,
                    "start": shift.get("start", ""),
                    "end": shift.get("end", ""),
                    "role": shift.get("role", ""),
                    "tier": tier,
                    "rate": rate,
                    "shift_type": shift.get("shift_type", "regular"),
                })

    fieldnames = ["employee_id", "name", "date", "start", "end", "role", "tier", "rate", "shift_type"]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return f"CSV exported to {filepath} ({len(rows)} shift records)"


def export_to_ical(schedule: Dict, employee_id: str, filepath: str) -> str:
    """
    Export individual employee schedule to iCal (.ics) format.
    Each shift = calendar event with location, role, rate in notes.
    Includes alerts: 1 hour before shift, 12 hours before shift.
    """
    employees = schedule.get("employees", [])
    facility = schedule.get("facility", "Main Facility")
    week_start_str = schedule.get("week_start", "2026-01-05")
    days_map = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5, "Sun": 6}

    try:
        week_start_date = datetime.strptime(week_start_str, "%Y-%m-%d")
    except ValueError:
        week_start_date = datetime(2026, 1, 5)

    # Find the employee
    emp = None
    for e in employees:
        if e.get("id") == employee_id:
            emp = e
            break

    if not emp:
        return f"Error: Employee {employee_id} not found in schedule"

    # Build iCal content
    cal_lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//WorkforceComplianceAI//ShiftPool//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{emp.get('name', 'Employee')} - Shifts",
    ]

    shifts = emp.get("shifts", {})
    for day, offset in days_map.items():
        shift = shifts.get(day, {})
        if not shift:
            continue

        shift_date = week_start_date + timedelta(days=offset)
        start_time = shift.get("start", "06:00")
        end_time = shift.get("end", "14:00")
        role = shift.get("role", "General")
        rate = emp.get("rate", 25.0)
        shift_type = shift.get("shift_type", "regular")

        # Parse times
        start_h, start_m = _parse_time(start_time)
        end_h, end_m = _parse_time(end_time)

        dtstart = shift_date.replace(hour=start_h, minute=start_m)
        dtend = shift_date.replace(hour=end_h, minute=end_m)

        uid = str(uuid.uuid4())

        cal_lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTART:{dtstart.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{dtend.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:Shift - {role}",
            f"LOCATION:{facility}",
            f"DESCRIPTION:Role: {role}\\nRate: ${rate}/hr\\nType: {shift_type}\\nTier: {emp.get('tier', 1)}",
            # 12 hours before alarm
            "BEGIN:VALARM",
            "TRIGGER:-PT12H",
            "ACTION:DISPLAY",
            "DESCRIPTION:Shift tomorrow",
            "END:VALARM",
            # 1 hour before alarm
            "BEGIN:VALARM",
            "TRIGGER:-PT1H",
            "ACTION:DISPLAY",
            "DESCRIPTION:Shift starts in 1 hour",
            "END:VALARM",
            "END:VEVENT",
        ])

    cal_lines.append("END:VCALENDAR")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\r\n".join(cal_lines))

    return f"iCal exported to {filepath} for {emp.get('name', employee_id)}"


# =============================================================================
# 2. PRINT FORMATS
# =============================================================================

def generate_breakroom_poster(schedule: Dict) -> str:
    """
    Generate a large-format breakroom poster for wall posting.
    Big, readable grid showing all employees, all shifts, all days.
    Highlights changes from last week in yellow (marked with *).
    """
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    employees = schedule.get("employees", [])
    changes = schedule.get("changes_from_last_week", [])
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    # Build change lookup
    change_set = set()
    for change in changes:
        change_set.add((change.get("employee_id", ""), change.get("day", "")))

    poster_width = 110
    lines = []
    lines.append("+" + "=" * (poster_width - 2) + "+")
    lines.append("|" + f" {facility} - WEEKLY SCHEDULE ".center(poster_width - 2) + "|")
    lines.append("|" + f" Week of {week_start} ".center(poster_width - 2) + "|")
    lines.append("+" + "=" * (poster_width - 2) + "+")
    lines.append("")

    # Table header
    col_name = 20
    col_day = 12
    header = f"| {'EMPLOYEE':<{col_name}}"
    for day in days:
        header += f"| {day:^{col_day}}"
    header += "|"
    lines.append(header)
    lines.append("|" + "-" * (col_name + 1) + ("+" + "-" * (col_day + 1)) * 7 + "|")

    # Employee rows
    for emp in employees:
        name = emp.get("name", "Unknown")[:col_name]
        emp_id = emp.get("id", "")
        shifts = emp.get("shifts", {})
        row = f"| {name:<{col_name}}"
        for day in days:
            shift = shifts.get(day, {})
            if shift:
                cell_text = f"{shift.get('start', '')}-{shift.get('end', '')}"
                # Mark changes with asterisk
                if (emp_id, day) in change_set:
                    cell_text = f"*{cell_text}*"
            else:
                cell_text = "OFF"
            row += f"| {cell_text:^{col_day}}"
        row += "|"
        lines.append(row)

    lines.append("|" + "-" * (col_name + 1) + ("+" + "-" * (col_day + 1)) * 7 + "|")
    lines.append("")
    lines.append("  * = Changed from last week")
    lines.append("")
    posted_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines.append(f'  Schedule posted {posted_date}. Questions? Text ShiftPool.')
    lines.append("")
    lines.append("+" + "=" * (poster_width - 2) + "+")

    return "\n".join(lines)


def generate_individual_card(schedule: Dict, employee_id: str) -> str:
    """
    Generate a per-person printout (wallet-sized / half-page).
    Shows just their shifts for the week.
    Includes QR code placeholder and emergency contact info.
    """
    employees = schedule.get("employees", [])
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    manager_phone = schedule.get("manager_phone", "(555) 123-4567")
    app_url = schedule.get("app_url", "https://shiftpool.app")

    emp = None
    for e in employees:
        if e.get("id") == employee_id:
            emp = e
            break

    if not emp:
        return f"Error: Employee {employee_id} not found"

    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    shifts = emp.get("shifts", {})

    lines = []
    lines.append("+" + "-" * 44 + "+")
    lines.append(f"| {'SHIFTPOOL - YOUR WEEKLY SCHEDULE':^42} |")
    lines.append("+" + "-" * 44 + "+")
    lines.append(f"| Name: {emp.get('name', 'Unknown'):<36} |")
    lines.append(f"| Week: {week_start:<36} |")
    lines.append(f"| Facility: {facility:<32} |")
    lines.append("|" + "-" * 44 + "|")

    for day in days:
        shift = shifts.get(day, {})
        if shift:
            time_str = f"{shift.get('start', '')}-{shift.get('end', '')}"
            role = shift.get("role", "")
            lines.append(f"|  {day}: {time_str:<12} ({role}){' ' * (20 - len(role))}|")
        else:
            lines.append(f"|  {day}: {'OFF':<35} |")

    lines.append("|" + "-" * 44 + "|")
    lines.append(f"| QR: {app_url}/s/{employee_id[:8]:<27} |")
    lines.append("|  [Scan for real-time updates]              |")
    lines.append("|" + "-" * 44 + "|")
    lines.append(f"| Manager: {manager_phone:<33} |")
    lines.append(f"| Emergency: Call (555) 911-0000{' ' * 13}|")
    lines.append("+" + "-" * 44 + "+")

    return "\n".join(lines)


# =============================================================================
# 3. SHARING
# =============================================================================

def generate_share_link(schedule: Dict, expiry_hours: int = 168) -> Dict:
    """
    Generate a unique, token-based shareable URL for the schedule.
    Configurable expiry (default 1 week = 168 hours).
    Read-only view.
    """
    token = secrets.token_urlsafe(32)
    schedule_id = schedule.get("id", str(uuid.uuid4()))
    base_url = schedule.get("app_url", "https://shiftpool.app")
    created_at = datetime.now()
    expires_at = created_at + timedelta(hours=expiry_hours)

    share_data = {
        "url": f"{base_url}/shared/{token}",
        "token": token,
        "schedule_id": schedule_id,
        "created_at": created_at.isoformat(),
        "expires_at": expires_at.isoformat(),
        "expiry_hours": expiry_hours,
        "access": "read-only",
        "views": {
            "full": f"{base_url}/shared/{token}?view=full",
            "individual": f"{base_url}/shared/{token}?view=individual&emp={{employee_id}}",
        },
    }

    return share_data


def email_schedule(schedule: Dict, recipients: List[str], format: str = "pdf",
                   smtp_config: Optional[Dict] = None) -> Dict:
    """
    Email the schedule to all employees or a specific group.
    Attach PDF or include inline HTML table.
    Returns send status (does not actually send without SMTP config).
    """
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    app_url = schedule.get("app_url", "https://shiftpool.app")

    subject = f"Your Schedule for Week of {week_start} - {facility}"

    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <h2>Your schedule for the week of {week_start} is ready.</h2>
        <p>Facility: {facility}</p>
        <p>
            <a href="{app_url}/schedule" style="background:#2196F3;color:white;padding:10px 20px;text-decoration:none;border-radius:4px;">View in App</a>
            &nbsp;
            <a href="{app_url}/schedule/download" style="background:#4CAF50;color:white;padding:10px 20px;text-decoration:none;border-radius:4px;">Download PDF</a>
        </p>
        {_generate_html_table(schedule)}
        <p style="color:#666;font-size:12px;">This is an automated message from ShiftPool. Do not reply.</p>
    </body>
    </html>
    """

    result = {
        "recipients": recipients,
        "subject": subject,
        "format": format,
        "status": "prepared",
        "total_recipients": len(recipients),
        "message_preview": f"Your schedule for {week_start} is ready. [View in app] [Download PDF]",
    }

    if smtp_config:
        # Attempt actual send
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = smtp_config.get("from_address", "schedule@shiftpool.app")
            msg["To"] = ", ".join(recipients)
            msg.attach(MIMEText(html_body, "html"))

            server = smtplib.SMTP(smtp_config.get("host", "localhost"), smtp_config.get("port", 587))
            server.starttls()
            server.login(smtp_config.get("username", ""), smtp_config.get("password", ""))
            server.sendmail(msg["From"], recipients, msg.as_string())
            server.quit()
            result["status"] = "sent"
        except Exception as e:
            result["status"] = f"failed: {str(e)}"
    else:
        result["status"] = "prepared (no SMTP config - would send in production)"

    return result


def sms_schedule_summary(schedule: Dict, employee_id: str) -> str:
    """
    Generate an ultra-brief SMS schedule summary for an employee.
    Format: "Your shifts next week: Mon 6am, Tue 6am, ... Reply OPEN to see options."
    """
    employees = schedule.get("employees", [])
    days_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    emp = None
    for e in employees:
        if e.get("id") == employee_id:
            emp = e
            break

    if not emp:
        return f"Error: Employee {employee_id} not found"

    shifts = emp.get("shifts", {})
    parts = []
    surge_days = []

    for day in days_order:
        shift = shifts.get(day, {})
        if shift:
            start = shift.get("start", "")
            # Abbreviate time (e.g., "06:00" -> "6am", "14:00" -> "2pm")
            parts.append(f"{day} {_abbreviate_time(start)}")
        else:
            parts.append(f"{day} OFF")
            surge_days.append(day)

    shift_text = ", ".join(parts)
    message = f"Your shifts next week: {shift_text}."

    if surge_days:
        message += f" Surge available {surge_days[0]}."

    message += " Reply OPEN to see options."

    return message


def slack_post(schedule: Dict, channel: str) -> Dict:
    """
    Generate a formatted Slack block message for posting to a team channel.
    Includes schedule table and action buttons.
    """
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    employees = schedule.get("employees", [])
    app_url = schedule.get("app_url", "https://shiftpool.app")

    # Build a simple text table for Slack
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    table_lines = [f"{'Employee':<15}" + "".join(f"{d:<8}" for d in days)]
    table_lines.append("-" * 71)

    for emp in employees[:15]:  # Limit for Slack message size
        name = emp.get("name", "Unknown")[:13]
        shifts = emp.get("shifts", {})
        row = f"{name:<15}"
        for day in days:
            shift = shifts.get(day, {})
            if shift:
                row += f"{shift.get('start', ''):<8}"
            else:
                row += f"{'OFF':<8}"
        table_lines.append(row)

    schedule_text = "\n".join(table_lines)

    # Slack Block Kit format
    slack_message = {
        "channel": channel,
        "blocks": [
            {
                "type": "header",
                "text": {
                    "type": "plain_text",
                    "text": f"{facility} - Schedule for {week_start}",
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"```\n{schedule_text}\n```",
                }
            },
            {
                "type": "actions",
                "elements": [
                    {
                        "type": "button",
                        "text": {"type": "plain_text", "text": "View Full Schedule"},
                        "url": f"{app_url}/schedule",
                        "style": "primary",
                    },
                    {
                        "type": "button",
                        "text": {"type": "plain_text", "text": "Claim Open Shifts"},
                        "url": f"{app_url}/shifts/open",
                    },
                ]
            },
            {
                "type": "context",
                "elements": [
                    {
                        "type": "mrkdwn",
                        "text": f"Posted by ShiftPool | {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    }
                ]
            },
        ],
    }

    return slack_message


# =============================================================================
# 4. IMPORT
# =============================================================================

def import_from_csv(filepath: str) -> Dict:
    """
    Import schedule from external CSV file.
    Auto-detects column mapping (flexible column names).
    Validates required fields (employee, date, start, end).
    Returns standardized schedule dict.
    """
    # Column name mappings (flexible matching)
    column_aliases = {
        "employee_id": ["employee_id", "emp_id", "id", "employee_number", "emp_no", "worker_id"],
        "name": ["name", "employee_name", "full_name", "worker_name", "emp_name"],
        "date": ["date", "shift_date", "work_date", "schedule_date", "day"],
        "start": ["start", "start_time", "shift_start", "clock_in", "begin"],
        "end": ["end", "end_time", "shift_end", "clock_out", "finish"],
        "role": ["role", "position", "job_title", "job", "department"],
        "tier": ["tier", "level", "grade", "classification"],
        "rate": ["rate", "pay_rate", "hourly_rate", "wage"],
        "shift_type": ["shift_type", "type", "category", "shift_category"],
    }

    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        raw_headers = reader.fieldnames or []

        # Auto-detect column mapping
        mapping = {}
        for standard_name, aliases in column_aliases.items():
            for header in raw_headers:
                if header.lower().strip() in aliases:
                    mapping[standard_name] = header
                    break

        # Validate required fields
        required = ["date", "start", "end"]
        missing = [r for r in required if r not in mapping]
        if missing:
            return {
                "error": f"Missing required columns: {missing}",
                "detected_headers": raw_headers,
                "mapping": mapping,
            }

        # Need at least employee_id or name
        if "employee_id" not in mapping and "name" not in mapping:
            return {
                "error": "Need at least employee_id or name column",
                "detected_headers": raw_headers,
                "mapping": mapping,
            }

        # Parse rows
        rows = []
        for row in reader:
            parsed = {}
            for standard_name, original_header in mapping.items():
                parsed[standard_name] = row.get(original_header, "").strip()
            rows.append(parsed)

    # Convert to standardized schedule format
    employees_dict = {}
    for row in rows:
        emp_key = row.get("employee_id") or row.get("name", "Unknown")
        if emp_key not in employees_dict:
            employees_dict[emp_key] = {
                "id": row.get("employee_id", emp_key),
                "name": row.get("name", emp_key),
                "tier": int(row.get("tier", 1)) if row.get("tier") else 1,
                "rate": float(row.get("rate", 25.0)) if row.get("rate") else 25.0,
                "role": row.get("role", "General"),
                "shifts": {},
            }

        # Map date to day of week
        try:
            date_obj = datetime.strptime(row["date"], "%Y-%m-%d")
            day_name = date_obj.strftime("%a")  # Mon, Tue, etc.
        except ValueError:
            day_name = row["date"]

        employees_dict[emp_key]["shifts"][day_name] = {
            "start": row.get("start", ""),
            "end": row.get("end", ""),
            "role": row.get("role", "General"),
            "shift_type": row.get("shift_type", "regular"),
        }

    schedule = {
        "employees": list(employees_dict.values()),
        "import_source": filepath,
        "import_date": datetime.now().isoformat(),
        "total_records": len(rows),
        "column_mapping": mapping,
    }

    return schedule


def import_from_excel(filepath: str) -> Dict:
    """
    Import schedule from Excel file.
    Supports multiple sheet formats.
    Auto-detects header row.
    Handles merged cells and date formats.
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    if not os.path.exists(filepath):
        return {"error": f"File not found: {filepath}"}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Auto-detect header row (first row with multiple non-empty cells)
    header_row = 1
    for row_idx in range(1, min(10, ws.max_row + 1)):
        non_empty = sum(1 for col in range(1, ws.max_column + 1)
                        if ws.cell(row=row_idx, column=col).value is not None)
        if non_empty >= 3:
            header_row = row_idx
            break

    # Read headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        headers.append(str(val).strip().lower() if val else f"col_{col}")

    # Column aliases (same as CSV)
    column_aliases = {
        "employee_id": ["employee_id", "emp_id", "id", "employee_number", "emp_no", "worker_id"],
        "name": ["name", "employee_name", "full_name", "worker_name", "emp_name", "employee"],
        "date": ["date", "shift_date", "work_date", "schedule_date", "day"],
        "start": ["start", "start_time", "shift_start", "clock_in", "begin", "start time"],
        "end": ["end", "end_time", "shift_end", "clock_out", "finish", "end time"],
        "role": ["role", "position", "job_title", "job", "department"],
        "tier": ["tier", "level", "grade", "classification"],
        "rate": ["rate", "pay_rate", "hourly_rate", "wage"],
    }

    mapping = {}
    for standard_name, aliases in column_aliases.items():
        for idx, header in enumerate(headers):
            if header in aliases:
                mapping[standard_name] = idx
                break

    # Parse data rows
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        for standard_name, col_idx in mapping.items():
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if cell_value is not None:
                # Handle date objects
                if isinstance(cell_value, datetime):
                    if standard_name == "date":
                        cell_value = cell_value.strftime("%Y-%m-%d")
                    else:
                        cell_value = cell_value.strftime("%H:%M")
                row_data[standard_name] = str(cell_value).strip()
            else:
                row_data[standard_name] = ""
        if any(row_data.values()):
            rows.append(row_data)

    # Convert to schedule format (same logic as CSV)
    employees_dict = {}
    for row in rows:
        emp_key = row.get("employee_id") or row.get("name", "Unknown")
        if not emp_key:
            continue
        if emp_key not in employees_dict:
            employees_dict[emp_key] = {
                "id": row.get("employee_id", emp_key),
                "name": row.get("name", emp_key),
                "tier": int(row.get("tier", 1)) if row.get("tier") else 1,
                "rate": float(row.get("rate", 25.0)) if row.get("rate") else 25.0,
                "role": row.get("role", "General"),
                "shifts": {},
            }

        date_str = row.get("date", "")
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            day_name = date_obj.strftime("%a")
        except ValueError:
            day_name = date_str

        if row.get("start"):
            employees_dict[emp_key]["shifts"][day_name] = {
                "start": row.get("start", ""),
                "end": row.get("end", ""),
                "role": row.get("role", "General"),
                "shift_type": "regular",
            }

    wb.close()

    return {
        "employees": list(employees_dict.values()),
        "import_source": filepath,
        "import_date": datetime.now().isoformat(),
        "total_records": len(rows),
        "sheets_available": wb.sheetnames if hasattr(wb, "sheetnames") else [],
    }


def import_from_kronos(api_response: Dict) -> Dict:
    """
    Parse UKG/Kronos schedule export format.
    Converts standard Kronos API response to internal schedule format.
    """
    schedules = api_response.get("scheduleItems", api_response.get("data", []))
    employees_dict = {}

    for item in schedules:
        emp_id = item.get("employeeId", item.get("personNumber", ""))
        emp_name = item.get("employeeName", item.get("personFullName", "Unknown"))
        date_str = item.get("date", item.get("startDate", ""))
        start_time = item.get("startTime", item.get("inPunch", ""))
        end_time = item.get("endTime", item.get("outPunch", ""))
        job = item.get("jobTitle", item.get("laborCategory", "General"))

        if emp_id not in employees_dict:
            employees_dict[emp_id] = {
                "id": emp_id,
                "name": emp_name,
                "tier": 1,
                "rate": 25.0,
                "role": job,
                "shifts": {},
            }

        try:
            date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
            day_name = date_obj.strftime("%a")
        except (ValueError, TypeError):
            day_name = date_str

        employees_dict[emp_id]["shifts"][day_name] = {
            "start": start_time,
            "end": end_time,
            "role": job,
            "shift_type": item.get("scheduleType", "regular"),
        }

    return {
        "employees": list(employees_dict.values()),
        "import_source": "kronos_api",
        "import_date": datetime.now().isoformat(),
        "total_records": len(schedules),
    }


def import_from_adp(api_response: Dict) -> Dict:
    """
    Parse ADP workforce schedule format.
    Converts ADP API response to internal schedule format.
    """
    workers = api_response.get("workers", api_response.get("scheduleEntries", []))
    employees_dict = {}

    for worker in workers:
        emp_id = worker.get("associateOID", worker.get("workerId", ""))
        name_data = worker.get("person", {}).get("legalName", {})
        emp_name = f"{name_data.get('givenName', '')} {name_data.get('familyName', '')}".strip()
        if not emp_name:
            emp_name = worker.get("workerName", worker.get("name", "Unknown"))

        schedules = worker.get("schedules", worker.get("shifts", []))

        if emp_id not in employees_dict:
            employees_dict[emp_id] = {
                "id": emp_id,
                "name": emp_name,
                "tier": 1,
                "rate": float(worker.get("hourlyRate", 25.0)),
                "role": worker.get("jobTitle", worker.get("position", "General")),
                "shifts": {},
            }

        for sched in schedules:
            date_str = sched.get("date", sched.get("scheduledDate", ""))
            start_time = sched.get("startTime", sched.get("scheduledStartTime", ""))
            end_time = sched.get("endTime", sched.get("scheduledEndTime", ""))

            try:
                date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
                day_name = date_obj.strftime("%a")
            except (ValueError, TypeError):
                day_name = date_str

            employees_dict[emp_id]["shifts"][day_name] = {
                "start": start_time,
                "end": end_time,
                "role": employees_dict[emp_id]["role"],
                "shift_type": sched.get("shiftType", "regular"),
            }

    return {
        "employees": list(employees_dict.values()),
        "import_source": "adp_api",
        "import_date": datetime.now().isoformat(),
        "total_records": len(workers),
    }


# =============================================================================
# 5. REPORT GENERATION
# =============================================================================

def weekly_summary_report(schedule: Dict, pool_data: Dict, compliance_data: Dict) -> str:
    """
    Generate a 1-page weekly manager report.
    KPIs: fill rate, cost, compliance score, trends.
    Auto-generated format for Monday morning delivery.
    """
    facility = schedule.get("facility", "Main Facility")
    week_start = schedule.get("week_start", "2026-01-05")
    employees = schedule.get("employees", [])

    fill_rate = pool_data.get("fill_rate", 92)
    open_shifts = pool_data.get("open_shifts", 3)
    total_shifts = pool_data.get("total_shifts", 140)
    surge_available = pool_data.get("surge_available", 8)

    compliance_score = compliance_data.get("score", 98)
    violations = compliance_data.get("violations", [])
    certifications_expiring = compliance_data.get("certifications_expiring", 0)

    cost_data = schedule.get("cost_summary", {})
    total_cost = cost_data.get("total_pool_cost", 45000)
    savings = cost_data.get("savings", 8500)
    savings_pct = cost_data.get("savings_pct", 15.9)

    report = []
    report.append("=" * 70)
    report.append(f"  WEEKLY SCHEDULE SUMMARY - {facility}".center(70))
    report.append(f"  Week of {week_start}".center(70))
    report.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}".center(70))
    report.append("=" * 70)
    report.append("")

    # KPIs
    report.append("  KEY PERFORMANCE INDICATORS")
    report.append("  " + "-" * 40)
    report.append(f"  Fill Rate:          {fill_rate}% ({total_shifts - open_shifts}/{total_shifts} shifts)")
    report.append(f"  Open Shifts:        {open_shifts}")
    report.append(f"  Surge Pool Ready:   {surge_available} employees")
    report.append(f"  Total Headcount:    {len(employees)}")
    report.append("")

    # Cost
    report.append("  COST SUMMARY")
    report.append("  " + "-" * 40)
    report.append(f"  Total Labor Cost:   ${total_cost:,.2f}")
    report.append(f"  Savings vs Trad.:   ${savings:,.2f} ({savings_pct:.1f}%)")
    report.append("")

    # Compliance
    report.append("  COMPLIANCE")
    report.append("  " + "-" * 40)
    report.append(f"  Score:              {compliance_score}/100")
    report.append(f"  Violations:         {len(violations)}")
    report.append(f"  Certs Expiring:     {certifications_expiring}")
    if violations:
        for v in violations[:3]:
            report.append(f"    - {v.get('type', 'Unknown')}: {v.get('details', '')}")
    report.append("")

    # Trends
    report.append("  TRENDS (vs last week)")
    report.append("  " + "-" * 40)
    trend_fill = pool_data.get("fill_rate_trend", "+2%")
    trend_cost = cost_data.get("cost_trend", "-3%")
    trend_compliance = compliance_data.get("score_trend", "+1")
    report.append(f"  Fill Rate:          {trend_fill}")
    report.append(f"  Cost:               {trend_cost}")
    report.append(f"  Compliance:         {trend_compliance}")
    report.append("")
    report.append("=" * 70)

    return "\n".join(report)


def monthly_executive_report(data: Dict) -> str:
    """
    Generate a multi-page monthly executive report for CFO/VP.
    Covers: ROI, cost savings, compliance improvement, employee satisfaction proxy.
    Charts noted as placeholders for where real charts would go.
    """
    facility = data.get("facility", "Main Facility")
    month = data.get("month", "January 2026")
    roi_data = data.get("roi", {})
    cost_data = data.get("costs", {})
    compliance_data = data.get("compliance", {})
    satisfaction_data = data.get("satisfaction", {})

    report = []
    report.append("=" * 70)
    report.append(f"  MONTHLY EXECUTIVE REPORT".center(70))
    report.append(f"  {facility} - {month}".center(70))
    report.append(f"  Prepared: {datetime.now().strftime('%Y-%m-%d')}".center(70))
    report.append("=" * 70)
    report.append("")

    # Page 1: ROI Summary
    report.append("  1. RETURN ON INVESTMENT")
    report.append("  " + "=" * 50)
    report.append(f"  Platform Investment:     ${roi_data.get('investment', 50000):,.2f}")
    report.append(f"  Monthly Savings:         ${roi_data.get('monthly_savings', 34000):,.2f}")
    report.append(f"  Cumulative Savings:      ${roi_data.get('cumulative_savings', 102000):,.2f}")
    report.append(f"  ROI:                     {roi_data.get('roi_pct', 204):.0f}%")
    report.append(f"  Payback Period:          {roi_data.get('payback_months', 1.5):.1f} months")
    report.append("")
    report.append("  [CHART: Monthly savings trend - bar chart would go here]")
    report.append("  [CHART: Cumulative ROI curve - line chart would go here]")
    report.append("")

    # Page 2: Cost Analysis
    report.append("  2. COST ANALYSIS")
    report.append("  " + "=" * 50)
    report.append(f"  Total Labor Cost:        ${cost_data.get('total_labor', 180000):,.2f}")
    report.append(f"  Traditional Equivalent:  ${cost_data.get('traditional_equivalent', 214000):,.2f}")
    report.append(f"  Net Savings:             ${cost_data.get('net_savings', 34000):,.2f}")
    report.append(f"  Overtime Reduction:      {cost_data.get('overtime_reduction_pct', 45):.0f}%")
    report.append(f"  Agency Spend Eliminated: ${cost_data.get('agency_eliminated', 22000):,.2f}")
    report.append("")
    report.append("  Cost Breakdown:")
    report.append(f"    Tier 1 (base):         ${cost_data.get('tier1_cost', 120000):,.2f} ({cost_data.get('tier1_pct', 67):.0f}%)")
    report.append(f"    Tier 2 (+15%):         ${cost_data.get('tier2_cost', 42000):,.2f} ({cost_data.get('tier2_pct', 23):.0f}%)")
    report.append(f"    Tier 3 (+25%):         ${cost_data.get('tier3_cost', 18000):,.2f} ({cost_data.get('tier3_pct', 10):.0f}%)")
    report.append("")
    report.append("  [CHART: Cost breakdown pie chart would go here]")
    report.append("  [CHART: Month-over-month cost comparison would go here]")
    report.append("")

    # Page 3: Compliance
    report.append("  3. COMPLIANCE IMPROVEMENT")
    report.append("  " + "=" * 50)
    report.append(f"  Current Score:           {compliance_data.get('current_score', 97)}/100")
    report.append(f"  Previous Month:          {compliance_data.get('previous_score', 94)}/100")
    report.append(f"  Improvement:             +{compliance_data.get('improvement', 3)} points")
    report.append(f"  Violations This Month:   {compliance_data.get('violations_count', 2)}")
    report.append(f"  Violations Last Month:   {compliance_data.get('previous_violations', 7)}")
    report.append(f"  Auto-Prevented:          {compliance_data.get('auto_prevented', 15)} potential violations")
    report.append("")
    report.append("  Top Compliance Areas:")
    for area in compliance_data.get("areas", [{"name": "Overtime Limits", "score": 100},
                                               {"name": "Rest Periods", "score": 98},
                                               {"name": "Certification", "score": 95}]):
        report.append(f"    - {area['name']}: {area['score']}/100")
    report.append("")
    report.append("  [CHART: Compliance score trend - line chart would go here]")
    report.append("")

    # Page 4: Employee Satisfaction Proxy
    report.append("  4. EMPLOYEE SATISFACTION INDICATORS")
    report.append("  " + "=" * 50)
    report.append(f"  Schedule Preference Match: {satisfaction_data.get('preference_match', 89):.0f}%")
    report.append(f"  Shift Swap Success Rate:   {satisfaction_data.get('swap_success', 94):.0f}%")
    report.append(f"  Voluntary Surge Uptake:    {satisfaction_data.get('surge_uptake', 72):.0f}%")
    report.append(f"  No-Show Rate:              {satisfaction_data.get('no_show_rate', 1.2):.1f}%")
    report.append(f"  Turnover Rate (monthly):   {satisfaction_data.get('turnover_rate', 2.1):.1f}%")
    report.append(f"  App Engagement:            {satisfaction_data.get('app_engagement', 85):.0f}%")
    report.append("")
    report.append("  [CHART: Employee satisfaction proxy trends would go here]")
    report.append("")

    report.append("=" * 70)
    report.append("  END OF REPORT")
    report.append("=" * 70)

    return "\n".join(report)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _calculate_headcount(schedule: Dict) -> Dict:
    """Calculate headcount per day from schedule."""
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    headcount = {day: 0 for day in days}
    for emp in schedule.get("employees", []):
        shifts = emp.get("shifts", {})
        for day in days:
            if shifts.get(day):
                headcount[day] += 1
    return headcount


def _parse_time(time_str: str) -> tuple:
    """Parse time string to (hour, minute) tuple."""
    try:
        if ":" in time_str:
            parts = time_str.replace("am", "").replace("pm", "").split(":")
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            if "pm" in time_str.lower() and h != 12:
                h += 12
            elif "am" in time_str.lower() and h == 12:
                h = 0
            return (h, m)
        else:
            h = int(time_str.replace("am", "").replace("pm", ""))
            if "pm" in time_str.lower() and h != 12:
                h += 12
            return (h, 0)
    except (ValueError, IndexError):
        return (0, 0)


def _abbreviate_time(time_str: str) -> str:
    """Convert time like '06:00' to '6am', '14:00' to '2pm'."""
    try:
        h, m = _parse_time(time_str)
        if h == 0:
            return "12am"
        elif h < 12:
            return f"{h}am"
        elif h == 12:
            return "12pm"
        else:
            return f"{h - 12}pm"
    except (ValueError, IndexError):
        return time_str


def _generate_html_table(schedule: Dict) -> str:
    """Generate an HTML table for email embedding."""
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    employees = schedule.get("employees", [])

    tier_colors = {1: "#ADD8E6", 2: "#90EE90", 3: "#FFC882"}

    html = '<table style="border-collapse:collapse;width:100%;font-size:12px;">'
    html += "<tr><th style='border:1px solid #ddd;padding:6px;'>Employee</th>"
    for day in days:
        html += f"<th style='border:1px solid #ddd;padding:6px;'>{day}</th>"
    html += "</tr>"

    for emp in employees[:20]:
        tier = emp.get("tier", 1)
        html += "<tr>"
        html += f"<td style='border:1px solid #ddd;padding:4px;'>{emp.get('name', '')}</td>"
        shifts = emp.get("shifts", {})
        for day in days:
            shift = shifts.get(day, {})
            if shift:
                bg = tier_colors.get(tier, "#fff")
                text = f"{shift.get('start', '')}-{shift.get('end', '')}"
                html += f"<td style='border:1px solid #ddd;padding:4px;background:{bg};text-align:center;'>{text}</td>"
            else:
                html += "<td style='border:1px solid #ddd;padding:4px;text-align:center;color:#999;'>OFF</td>"
        html += "</tr>"

    html += "</table>"
    return html


def _generate_sample_schedule(num_employees: int = 20) -> Dict:
    """Generate a sample schedule for demonstration purposes."""
    import random

    roles = ["RN", "CNA", "LPN", "Charge RN", "Med Tech", "PCT", "Unit Clerk", "Supervisor"]
    first_names = ["Sarah", "Mike", "Jennifer", "David", "Lisa", "James", "Maria", "Robert",
                   "Emily", "Carlos", "Ashley", "Kevin", "Nicole", "Brian", "Amanda",
                   "Chris", "Jessica", "Daniel", "Michelle", "Tyler"]
    last_names = ["Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
                  "Rodriguez", "Martinez", "Anderson", "Taylor", "Thomas", "Moore", "Jackson",
                  "Martin", "Lee", "Perez", "Thompson", "White", "Harris"]

    shift_patterns = [
        {"start": "06:00", "end": "14:00"},
        {"start": "14:00", "end": "22:00"},
        {"start": "22:00", "end": "06:00"},
        {"start": "07:00", "end": "15:00"},
        {"start": "08:00", "end": "16:00"},
    ]

    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    employees = []

    for i in range(num_employees):
        tier = random.choice([1, 1, 1, 2, 2, 3])  # Weighted: more tier 1
        role = random.choice(roles)
        base_rate = {1: 25.0, 2: 28.75, 3: 31.25}[tier]
        reliability = round(random.uniform(0.85, 1.0), 2)

        # Generate shifts (5 days on, 2 off typically)
        work_days = random.sample(days, random.choice([4, 5, 5, 5, 6]))
        shift_pattern = random.choice(shift_patterns)

        shifts = {}
        for day in work_days:
            shifts[day] = {
                "start": shift_pattern["start"],
                "end": shift_pattern["end"],
                "role": role,
                "shift_type": "regular" if day not in ["Sat", "Sun"] else "weekend",
            }

        employees.append({
            "id": f"EMP{i + 1:04d}",
            "name": f"{first_names[i]} {last_names[i]}",
            "tier": tier,
            "rate": base_rate,
            "role": role,
            "reliability_score": reliability,
            "weekly_hours": len(work_days) * 8,
            "shifts": shifts,
        })

    schedule = {
        "id": str(uuid.uuid4()),
        "facility": "Sunrise Healthcare Center",
        "week_start": "2026-01-05",
        "week_end": "2026-01-11",
        "pool_status": "Active - 92% fill rate",
        "compliance_status": "Compliant - No violations",
        "manager_phone": "(555) 234-5678",
        "app_url": "https://shiftpool.app",
        "employees": employees,
        "violations": [],
        "changes_from_last_week": [
            {"employee_id": "EMP0003", "day": "Wed"},
            {"employee_id": "EMP0007", "day": "Fri"},
        ],
        "cost_summary": {
            "regular_pay": 36000.0,
            "overtime_premiums": 2400.0,
            "tier2_premiums": 3450.0,
            "tier3_premiums": 1875.0,
            "total_pool_cost": 43725.0,
            "traditional_cost": 52000.0,
            "savings": 8275.0,
            "savings_pct": 15.9,
            "cost_trend": "-3.2%",
        },
    }

    return schedule


# =============================================================================
# MAIN DEMONSTRATION
# =============================================================================

if __name__ == "__main__":
    import random
    random.seed(42)

    print("=" * 70)
    print("  WORKFORCE COMPLIANCE AI - Export & Sharing Module Demo")
    print("=" * 70)
    print()

    # Generate sample schedule for 20 employees
    print("[1] Generating sample schedule for 20 employees...")
    schedule = _generate_sample_schedule(20)
    print(f"    Created schedule for {len(schedule['employees'])} employees")
    print(f"    Facility: {schedule['facility']}")
    print(f"    Week: {schedule['week_start']} to {schedule['week_end']}")
    print()

    # Export to CSV
    print("[2] Exporting to CSV...")
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_schedule_export.csv")
    result = export_to_csv(schedule, csv_path)
    print(f"    {result}")
    print()

    # Generate breakroom poster
    print("[3] Generating breakroom poster (print format)...")
    poster = generate_breakroom_poster(schedule)
    print(poster)
    print()

    # Generate individual card
    print("[4] Generating individual schedule card for EMP0001...")
    card = generate_individual_card(schedule, "EMP0001")
    print(card)
    print()

    # Generate iCal content (show format)
    print("[5] Generating iCal (.ics) content for EMP0002...")
    ical_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "emp0002_schedule.ics")
    ical_result = export_to_ical(schedule, "EMP0002", ical_path)
    print(f"    {ical_result}")
    # Show a snippet of the .ics content
    if os.path.exists(ical_path):
        with open(ical_path, "r") as f:
            content = f.read()
        lines = content.split("\r\n")
        print("    --- iCal Preview (first 20 lines) ---")
        for line in lines[:20]:
            print(f"    {line}")
        print("    ...")
    print()

    # Share link generation
    print("[6] Generating shareable schedule link...")
    share = generate_share_link(schedule, expiry_hours=168)
    print(f"    URL: {share['url']}")
    print(f"    Expires: {share['expires_at']}")
    print(f"    Access: {share['access']}")
    print(f"    Full view: {share['views']['full']}")
    print()

    # SMS summary
    print("[7] Generating SMS schedule summary for EMP0001...")
    sms = sms_schedule_summary(schedule, "EMP0001")
    print(f"    Message: {sms}")
    print(f"    Characters: {len(sms)}")
    print()

    # Slack post preview
    print("[8] Generating Slack post for #scheduling channel...")
    slack = slack_post(schedule, "#scheduling")
    print(f"    Channel: {slack['channel']}")
    print(f"    Blocks: {len(slack['blocks'])} sections")
    print(f"    Header: {slack['blocks'][0]['text']['text']}")
    print()

    # Final summary
    print("=" * 70)
    print("  All export formats ready: PDF, Excel, CSV, iCal, SMS, Email, Slack, Print")
    print("=" * 70)
    print()
    print("  Supported Operations:")
    print("    EXPORT:  PDF | Excel (4 sheets) | CSV | iCal")
    print("    PRINT:   Breakroom Poster | Individual Card")
    print("    SHARE:   Link (expiring) | Email | SMS | Slack")
    print("    IMPORT:  CSV | Excel | Kronos/UKG | ADP")
    print("    REPORTS: Weekly Summary | Monthly Executive")
    print()
